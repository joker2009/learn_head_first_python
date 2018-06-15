import openpyxl
import os

import csv
# 将收到的数据写入到创建的文件中
def write_file(list):
    with open('2018-06-13.csv', 'a+', encoding='UTF-8') as f:

        for i in list:
            # print(i)
            f.write(i + '\n')
    print('文件写入成功，行数为')
    print(len(list))
    return len(list)

# 去掉None行和去重
def data_cleaning(data):
    all_data = []
    for i in data:
        # print(i[0:4])
        if i[0:4] != 'None':
            all_data.append(i)
    return all_data




# 在文件目录下寻找EXCEL文件
def find_file_path(path):
    file_paths = os.listdir(path)
    excel_path_list = []
    csv_path_list = []
    for names in file_paths:
        if names.endswith('.xlsx'):
            # 添加需要导入的数据包地址，采用/，并在末尾加一个/
            excel_path_list.append(path + '/' + names)
        # if names.endswith('.csv'):

    # print(path_list)
    return excel_path_list


# 从excel表中取数据
def excel_data(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[1]
    row_length = len(list(ws.rows))

    excel_data = []
    all_data = []
    for i in list(ws.columns)[5:12]:
        excel_data.append(i)
    for i in range(1,row_length):
        all_data.append(str(excel_data[0][i].value) + ',' + str(excel_data[1][i].value) + ',' + str(excel_data[6][i].value))
    return all_data


# 获得对应csv文件的某一列的所有值
def getCsvValue(filepath, colnum):
    # print filepath
    ret = []
    # print(filepath)
    reader = csv.reader(open(filepath, encoding='utf-8',mode='r'))
    # reader = csv.reader(filepath, 'rb')
    for line in reader:
        # print(line)
        # if reader.line_num == 1:
        #     # 跳过第一行属性名
        #     continue
        ret.append(line[colnum])
    return ret
# 遍历整个excel表的头部，寻找所需要的数据的表头
def excel_data_2(path):
    wb = openpyxl.load_workbook(path)
    row_length = len(list(wb.worksheets[0].rows))
    item_list = []
    for i in range(len(wb.worksheets)):  # 遍历每个sheet
        # print(i)
        ws = wb.worksheets[i]
        # row_length = len(list(ws.rows))
        for item in ws.iter_rows(max_col=12,max_row=8):

            # print(item[0].value)
            item_list.append(item)
    # print(item_list) # 返回一个二维数组
    # print(row_length)
    return item_list,row_length
# 查找表头是否有符合要求的数据，并给出对应的坐标
def search_data(list, length):
    all_data = []
    for i in range(len(list)):
        # pass
        # all_data.append(str(list[1][i].value)+str(list[2][i].value)+str(list[0][i].value))
        all_data.append(str(list[i].value))
        # 定位坐标
        if str(list[i].value) == '外箱扫描号码':
            print(list[i])
            print(str(list[i])[-3:-1])
        # if str(list[i].value) == '客户编码':
        # if str(list[i].value) == '客户名称':

    print(all_data)
    return all_data




if __name__ == '__main__':
    # excel_path = r"D:\learn_head_first_python\aveeno_weekly_data_process\Aveeno产品JB条码周报20180604-0610(1).xlsx"
    # c = input()
    # excel_path = r"C:\Users\joker\Desktop\20180612\Aveeno产品JB条码周报-上海补报201604-201605v1.xlsx"
    # a = excel_data(excel_path)
    # b = data_cleaning(a)
    # write_file(b)

    # print(find_file_path(c))

# 用于提取EXCEL数据并写入文件中，已完成，将excel表所在文件夹
#     c = r"C:\Users\joker\Desktop\20180614test"
#     c = eval(input('输入excel所在文件夹路径：'))
    c = input('输入excel所在文件夹路径：')
    path_list = find_file_path(c)
    d =0
    for i in range(len(path_list)):
        a = excel_data(path_list[i])
        b = data_cleaning(a)
        c = write_file(b)
        d =c+d
    # 打印总行数，确认数据数量
    print('文件总行数：')
    print(d)


# 查找CSV中一列数据，准备去重，待续
#     f = getCsvValue(r'D:\learn_head_first_python\aveeno_weekly_data_process\2018-06-13.csv', 2)
    # print(f)


# # 使用表头寻找方式查找，现在已经定位到坐标，未完待续
    # z = r"C:\Users\joker\Desktop\20180612\Aveeno产品JB条码周报20180604-0610(2).xlsx"
    # y = excel_data_2(z)
    # # print(y)
    # x =y[0]
    # # print(x)
    # # print(y[0][1])
    # # print(type(y[0][0]))
    # # print(y[0][0])
    # # print(type(y[1]))
    # print(y[1])
    # for i in range(len(x)):
    #     # print(x[i].value)
    #     search_data(x[i], y[1])

