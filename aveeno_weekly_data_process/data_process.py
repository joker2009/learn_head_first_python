import openpyxl
import csv
# from openpyxl import *
excel = openpyxl.load_workbook(r'D:\learn_head_first_python\aveeno_weekly_data_process\Aveeno产品JB条码周报20180604-0610(1).xlsx')
# a = excel.active
# b = excel.worksheets
# c =excel.encoding
# print(b[1])
b= excel.sheetnames
# print(b[1])
a = b[1]
print(excel[a])
ws = excel.worksheets[1]
# print(ws)
# for item in ws:
#     print(item)
# 生成器
# print(ws.rows)
row_length = len(list(ws.rows))
# print(list(ws.rows))
print(row_length)
data = []
# print(list(ws.columns)[1])
for i in list(ws.columns)[5:12]:
    data.append(i)
    # print(i)
# print(data)
print(data[0][0])
a = data[0][0]
b =data[1][0]
c= data[6][0]

print(a.value)
print(b.value)
print(c.value)
csv_data = []
for i in range(row_length):
    csv_data.append(str(data[0][i].value) +','+ str(data[1][i].value) +','+ str(data[6][i].value) )
print(csv_data)
print(type(csv_data))
print(csv_data[0])
# csv_file =[]
# for v in range(len(csv_data)):
#     csv_file.append(csv_data[v])
# print(csv_file)
print(csv_data[1698])
with open("test.csv","w",encoding='utf-8') as csvfile:
    # writer = csv.writer(csvfile)
    # # 先写入columns_name
    # for v in range(len(csv_data)):
    #     print(csv_data[v])
    #     writer.writerow(csv_data[v])
    for v in csv_data:
        if csv_data[v]:
            csvfile.write(v+'\n')

# 将list 写入到csv


# c = excel[a]
# print(c)
# print(excel.active)
# data = []
# for row in excel[a].iter_rows(min_row=2,max_row=4,min_col=2,max_col=4):
#     # data.append(row)
#     print(row)
#     print(row[0].value)



    # print(dir(row))
    # print(type(row))
# print(data)
# for i in data:
#     print(i)